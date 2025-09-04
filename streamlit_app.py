# streamlit_app.py

import streamlit as st
import pandas as pd
import io
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from st_aggrid import AgGrid, GridOptionsBuilder, JsCode
from st_aggrid.shared import GridUpdateMode

# --- PAGE CONFIG ---
st.set_page_config(
    page_title="BCT Data Availability",
    page_icon="📊",
    layout="wide",
)

# --- CUSTOM STYLING ---
st.markdown("""
    <style>
        body {
            background-color: #ffffff;
        }
        .reportview-container .main .block-container {
            padding-top: 2rem;
            padding-bottom: 2rem;
        }
        .stButton>button {
            background-color: #009999;
            color: white;
            border-radius: 8px;
            padding: 0.6em 1em;
        }
        .stDownloadButton>button {
            background-color: #0066cc;
            color: white;
            border-radius: 8px;
            padding: 0.6em 1em;
            font-weight: bold;
        }
        h1, h2, h3 {
            color: #004d66;
        }
        .custom-table thead tr {
            background-color: #004d66;
            color: white;
        }
        .custom-table td, .custom-table th {
            border: 1px solid #ccc;
            padding: 8px 12px;
        }
        .custom-table {
            border-collapse: collapse;
            width: 100%;
        }
    </style>
""", unsafe_allow_html=True)

# --- TITLE ---
st.title("📈 BCT Data Availability Dashboard")

# --- FILE UPLOAD ---
st.header("📂 Upload Required Files")

col1, col2 = st.columns(2)
with col1:
    master_file = st.file_uploader("Upload Master Excel File", type=["xlsx"])
with col2:
    uploaded_csvs = st.file_uploader(
        "Upload CSV Files",
        type=["csv"],
        accept_multiple_files=True
    )

# === PROCESSING ===
if master_file and uploaded_csvs:
    st.success("✅ Files uploaded successfully!")

    # --- READ MASTER FILE ---
    master_df = pd.read_excel(master_file, engine='openpyxl')
    master_df.columns = [col.strip().title() for col in master_df.columns]

    # --- READ CSV FILES ---
    all_data = []
    for file in uploaded_csvs:
        try:
            df = pd.read_csv(file, header=None, names=['Timestamp', 'Asset Name', 'Active Power', 'Wind Speed'], on_bad_lines='skip')
            df['Timestamp'] = pd.to_datetime(df['Timestamp'], dayfirst=True, errors='coerce')
            df['Date'] = df['Timestamp'].dt.date
            df = df.dropna(subset=['Timestamp', 'Asset Name'])
            all_data.append(df[['Timestamp', 'Date', 'Asset Name', 'Active Power']])
        except Exception as e:
            st.error(f"Error reading {file.name}: {e}")

    compiled_df = pd.concat(all_data, ignore_index=True)

    # === SHEET 1 ===
    sheet1 = compiled_df.merge(master_df, on='Asset Name', how='left')

    # === SHEET 2 ===
    sheet2_counts = compiled_df.groupby(['Asset Name', 'Date']).size().reset_index(name='Count')
    sheet2 = sheet2_counts.merge(master_df, on='Asset Name', how='left')
    sheet2 = sheet2.groupby(['Make', 'Site', 'Date'])['Count'].sum().reset_index()
    sheet2_pivot = sheet2.pivot(index=['Make', 'Site'], columns='Date', values='Count').fillna(0).astype(int)
    sheet2_pivot.columns = [col.strftime('%d-%m-%Y') for col in sheet2_pivot.columns]
    sheet2_pivot.reset_index(inplace=True)

    # === SHEET 3 ===
    status_rows = []
    all_dates = sorted(compiled_df['Date'].dropna().unique())
    for (make, site), group in master_df.groupby(['Make', 'Site']):
        assets = group['Asset Name'].tolist()
        for date in all_dates:
            date_data = compiled_df[(compiled_df['Asset Name'].isin(assets)) & (compiled_df['Date'] == date)]
            asset_counts = date_data.groupby('Asset Name').size()
            total_assets = len(assets)
            avg = asset_counts.sum() / total_assets if total_assets > 0 else 0
            status = "Data Available" if avg >= 130 else "Data Not Available"
            status_rows.append({'Make': make, 'Site': site, 'Date': date, 'Status': status})

    sheet3 = pd.DataFrame(status_rows)
    sheet3_pivot = sheet3.pivot(index=['Make', 'Site'], columns='Date', values='Status')
    sheet3_pivot.columns = [col.strftime('%d-%m-%Y') for col in sheet3_pivot.columns]
    sheet3_pivot.reset_index(inplace=True)

    # === EXPORT TO EXCEL ===
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        sheet1.to_excel(writer, index=False, sheet_name='Compiled Data')
        sheet2_pivot.to_excel(writer, index=False, sheet_name='Compiled Summary')
        sheet3_pivot.to_excel(writer, index=False, sheet_name='Result Data')

    # === COLOR SHEET 3 (EXCEL) ===
    output.seek(0)
    wb = load_workbook(output)
    ws = wb['Result Data']
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for row in ws.iter_rows(min_row=2, min_col=3):
        for cell in row:
            if cell.value == "Data Available":
                cell.fill = green_fill
            elif cell.value == "Data Not Available":
                cell.fill = red_fill

    final_output = io.BytesIO()
    wb.save(final_output)
    final_output.seek(0)

    # === DISPLAY FUNCTIONS ===
    def display_html_table(df, title):
        st.subheader(title)
        html = df.head(50).to_html(classes='custom-table', index=False, escape=False)
        st.markdown(html, unsafe_allow_html=True)

    def display_status_table(df):
        styled_df = df.copy()
        for col in styled_df.columns[2:]:
            styled_df[col] = styled_df[col].replace({
                'Data Available': '<span style="background-color:#c6efce; color:#006100; padding:4px; border-radius:4px;">Available</span>',
                'Data Not Available': '<span style="background-color:#ffc7ce; color:#9c0006; padding:4px; border-radius:4px;">Not Available</span>'
            })
        html = styled_df.to_html(escape=False, index=False, classes="custom-table")
        st.markdown(html, unsafe_allow_html=True)

    # === DISPLAY TABLES ===
    st.header("🔍 Preview of Processed Data")
    display_html_table(sheet1, "🗂 Compiled Data")
    display_html_table(sheet2_pivot, "📊 Compiled Summary")
    display_status_table(sheet3_pivot)

    # === DOWNLOAD BUTTON ===
    st.download_button(
        label="📥 Download Final Excel File",
        data=final_output,
        file_name=f"data_availability_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# === Streamlit UI ===

uploaded_files = st.file_uploader("Upload CSV files", accept_multiple_files=True, type='csv')

if uploaded_files:
    # Save uploaded files to temp dir (in-memory)
    tmp_files = []
    for uploaded_file in uploaded_files:
        tmp_files.append(uploaded_file)

    # Process files
    compiled_df, filtered_df, max_df, result_df = process_data(tmp_files)

    if compiled_df is not None:

        st.subheader("📊 Result Data with Flags (Interactive Grid with Filters)")

        # Add color column (for optional display logic)
        styled_result_df = result_df.copy()
        styled_result_df['Color'] = styled_result_df['TempSum'].apply(lambda x: 'red' if x > 1 else 'yellow' if x == 1 else 'green')

        # AgGrid configuration
        gb = GridOptionsBuilder.from_dataframe(styled_result_df)
        gb.configure_pagination(paginationAutoPageSize=True)
        gb.configure_default_column(
            wrapText=True,
            autoHeight=True,
            resizable=True,
            filter=True,
            sortable=True
        )

        # JavaScript code to color cells conditionally
        cell_style_jscode = JsCode("""
        function(params) {
            if (params.colDef.field == 'TempSum') {
                if (params.value > 1) {
                    return {backgroundColor: 'red', color: 'white', fontWeight: 'bold'};
                } else if (params.value == 1) {
                    return {backgroundColor: 'yellow', color: 'black', fontWeight: 'bold'};
                } else {
                    return {backgroundColor: 'green', color: 'white', fontWeight: 'bold'};
                }
            }

            if (params.colDef.field.startsWith('Temp') && params.colDef.field !== 'TempSum') {
                if (params.value == 1) {
                    return {backgroundColor: '#FF9999'};
                }
            }

            return {};
        }
        """)

        # Apply the JS style to columns
        gb.configure_columns(
            ['Temp11', 'Temp22', 'Temp33', 'Temp44', 'Temp55', 'Temp66', 'TempSum'],
            cellStyle=cell_style_jscode
        )

        gridOptions = gb.build()

        # Display interactive styled AgGrid
        AgGrid(
            styled_result_df,
            gridOptions=gridOptions,
            height=600,
            width='100%',
            theme="streamlit",  # Other options: "light", "dark", "blue"
            update_mode=GridUpdateMode.NO_UPDATE,
            fit_columns_on_grid_load=True
        )

        # Export Excel
        excel_buffer = create_excel(compiled_df, filtered_df, max_df, result_df)
        st.download_button(
            label="📥 Download Excel Report",
            data=excel_buffer,
            file_name="final_report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        # Show charts
        st.subheader("📈 Temperature Exceedance Charts")
        charts = plot_exceedance_charts(compiled_df)
        for asset, fig in charts.items():
            st.markdown(f"**{asset}**")
            st.pyplot(fig)

else:
    st.info("Please upload CSV files to begin processing.")
else:
    st.info("Please upload both Master Excel and at least one CSV file to continue.")

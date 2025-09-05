# streamlit_app.py

import streamlit as st
import pandas as pd
import io
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
import matplotlib.pyplot as plt
import zipfile
from io import BytesIO 
import plotly.express as px

# --- PAGE CONFIG ---
st.set_page_config(page_title="Multi-Process App", page_icon="🔧", layout="wide")

# --- SIDEBAR MENU ---
st.sidebar.title("🔧 Select Process")
process_choice = st.sidebar.radio("Choose a process to run:", ["📊 BCT Data Availability Dashboard", "⚙️ Temperature & Power Analysis"])

# --- PROCESS 1: Existing Dashboard ---
if process_choice == "📊 BCT Data Availability Dashboard":
    st.title("📈 BCT Data Availability Dashboard")
    
    # --- PAGE CONFIG ---


    # --- CUSTOM STYLING ---
    st.markdown("""
        <style>
            body {
                background-color: #ffffff;
            }
            .reportview-container .main .block-container {
                padding-top: 2rem;
                padding-bottom: 2rem;
            }
            .stButton>button {
                background-color: #009999;
                color: white;
                border-radius: 8px;
                padding: 0.6em 1em;
            }
            .stDownloadButton>button {
                background-color: #0066cc;
                color: white;
                border-radius: 8px;
                padding: 0.6em 1em;
                font-weight: bold;
            }
            h1, h2, h3 {
                color: #004d66;
            }
            .custom-table thead tr {
                background-color: #004d66;
                color: white;
            }
            .custom-table td, .custom-table th {
                border: 1px solid #ccc;
                padding: 8px 12px;
            }
            .custom-table {
                border-collapse: collapse;
                width: 100%;
            }
        </style>
    """, unsafe_allow_html=True)

    # --- TITLE ---
    
    # --- FILE UPLOAD ---
    st.header("📂 Upload Required Files")

    col1, col2 = st.columns(2)
    with col1:
        master_file = st.file_uploader("Upload Master Excel File", type=["xlsx"])
    with col2:
        uploaded_csvs = st.file_uploader(
            "Upload CSV Files",
            type=["csv"],
            accept_multiple_files=True
        )

    # === PROCESSING ===
    if master_file and uploaded_csvs:
        st.success("✅ Files uploaded successfully!")

        # --- READ MASTER FILE ---
        master_df = pd.read_excel(master_file, engine='openpyxl')
        master_df.columns = [col.strip().title() for col in master_df.columns]

        # --- READ CSV FILES ---
        all_data = []
        for file in uploaded_csvs:
            try:
                df = pd.read_csv(file, header=None, names=['Timestamp', 'Asset Name', 'Active Power', 'Wind Speed'], on_bad_lines='skip')
                df['Timestamp'] = pd.to_datetime(df['Timestamp'], dayfirst=True, errors='coerce')
                df['Date'] = df['Timestamp'].dt.date
                df = df.dropna(subset=['Timestamp', 'Asset Name'])
                all_data.append(df[['Timestamp', 'Date', 'Asset Name', 'Active Power']])
            except Exception as e:
                st.error(f"Error reading {file.name}: {e}")

        compiled_df = pd.concat(all_data, ignore_index=True)

        # === SHEET 1 ===
        sheet1 = compiled_df.merge(master_df, on='Asset Name', how='left')

        # === SHEET 2 ===
        sheet2_counts = compiled_df.groupby(['Asset Name', 'Date']).size().reset_index(name='Count')
        sheet2 = sheet2_counts.merge(master_df, on='Asset Name', how='left')
        sheet2 = sheet2.groupby(['Make', 'Site', 'Date'])['Count'].sum().reset_index()
        sheet2_pivot = sheet2.pivot(index=['Make', 'Site'], columns='Date', values='Count').fillna(0).astype(int)
        sheet2_pivot.columns = [col.strftime('%d-%m-%Y') for col in sheet2_pivot.columns]
        sheet2_pivot.reset_index(inplace=True)

        # === SHEET 3 ===
        status_rows = []
        all_dates = sorted(compiled_df['Date'].dropna().unique())
        for (make, site), group in master_df.groupby(['Make', 'Site']):
            assets = group['Asset Name'].tolist()
            for date in all_dates:
                date_data = compiled_df[(compiled_df['Asset Name'].isin(assets)) & (compiled_df['Date'] == date)]
                asset_counts = date_data.groupby('Asset Name').size()
                total_assets = len(assets)
                avg = asset_counts.sum() / total_assets if total_assets > 0 else 0
                status = "Data Available" if avg >= 130 else "Data Not Available"
                status_rows.append({'Make': make, 'Site': site, 'Date': date, 'Status': status})

        sheet3 = pd.DataFrame(status_rows)
        sheet3_pivot = sheet3.pivot(index=['Make', 'Site'], columns='Date', values='Status')
        sheet3_pivot.columns = [col.strftime('%d-%m-%Y') for col in sheet3_pivot.columns]
        sheet3_pivot.reset_index(inplace=True)

        # === EXPORT TO EXCEL ===
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            sheet1.to_excel(writer, index=False, sheet_name='Compiled Data')
            sheet2_pivot.to_excel(writer, index=False, sheet_name='Compiled Summary')
            sheet3_pivot.to_excel(writer, index=False, sheet_name='Result Data')

        # === COLOR SHEET 3 (EXCEL) ===
        output.seek(0)
        wb = load_workbook(output)
        ws = wb['Result Data']
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        for row in ws.iter_rows(min_row=2, min_col=3):
            for cell in row:
                if cell.value == "Data Available":
                    cell.fill = green_fill
                elif cell.value == "Data Not Available":
                    cell.fill = red_fill

        final_output = io.BytesIO()
        wb.save(final_output)
        final_output.seek(0)

        # === DISPLAY FUNCTIONS ===
        def display_html_table(df, title):
            st.subheader(title)
            html = df.head(50).to_html(classes='custom-table', index=False, escape=False)
            st.markdown(html, unsafe_allow_html=True)

        def display_status_table(df):
            styled_df = df.copy()
            for col in styled_df.columns[2:]:
                styled_df[col] = styled_df[col].replace({
                    'Data Available': '<span style="background-color:#c6efce; color:#006100; padding:4px; border-radius:4px;">Available</span>',
                    'Data Not Available': '<span style="background-color:#ffc7ce; color:#9c0006; padding:4px; border-radius:4px;">Not Available</span>'
                })
            html = styled_df.to_html(escape=False, index=False, classes="custom-table")
            st.markdown(html, unsafe_allow_html=True)

        # === DISPLAY TABLES ===
        st.header("🔍 Preview of Processed Data")
        display_status_table(sheet3_pivot)

        # === DOWNLOAD BUTTON ===
        st.download_button(
            label="📥 Download Final Excel File",
            data=final_output,
            file_name=f"data_availability_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    else:
        st.info("Please upload both Master Excel and at least one CSV file to continue.")


# --- PROCESS 2: Temperature & Power Analysis ---
elif process_choice == "⚙️ Temperature & Power Analysis":
    st.title("Temperature and Power Data Processor")

# === Constants ===
active_power_threshold = 500
temp_exceed_limit = 90

temp_columns = [
    'Temperaturemeasurementforgeneratorbearingdriveend',
    'Temperaturemeasurementforgeneratorbearingnondriveend',
    'GearboxHighSpeedShaftDrivenEndtemp',
    'GearboxHighSpeedShaftNonDrivenEndtemp',
    'MeasuredTemperatureofrotorbearing',
    'OilSumpTemp'
]

required_cols = temp_columns + ['Asset Name', 'ActivepowerGeneration', 'Date']

thresholds = {
    'Temperaturemeasurementforgeneratorbearingdriveend': 90,
    'Temperaturemeasurementforgeneratorbearingnondriveend': 90,
    'GearboxHighSpeedShaftDrivenEndtemp': 90,
    'GearboxHighSpeedShaftNonDrivenEndtemp': 90,
    'MeasuredTemperatureofrotorbearing': 60,
    'OilSumpTemp': 80,
}

def process_data(csv_files):
    raw_dfs = []
    for file in csv_files:
        try:
            df = pd.read_csv(file)
            if 'Date' in df.columns:
                df['Date'] = pd.to_datetime(df['Date'], format='%d-%m-%Y %H:%M:%S', errors='coerce')
            if not df.empty:
                raw_dfs.append(df)
        except Exception as e:
            st.warning(f"Error reading {file}: {e}")

    if not raw_dfs:
        st.error("No valid CSV files loaded.")
        return None, None, None, None

    compiled_df = pd.concat(raw_dfs, ignore_index=True)

    missing_cols = [col for col in required_cols if col not in compiled_df.columns]
    if missing_cols:
        st.error(f"Missing columns in data: {missing_cols}")
        return None, None, None, None

    filtered_df = compiled_df[(compiled_df['ActivepowerGeneration'] > 0)]

    max_df = filtered_df.groupby('Asset Name')[temp_columns + ['ActivepowerGeneration']].max().reset_index()

    result_df = max_df.copy()
    result_df['Temp11'] = (result_df[temp_columns[0]] > 90).astype(int)
    result_df['Temp22'] = (result_df[temp_columns[1]] > 90).astype(int)
    result_df['Temp33'] = (result_df[temp_columns[2]] > 90).astype(int)
    result_df['Temp44'] = (result_df[temp_columns[3]] > 90).astype(int)
    result_df['Temp55'] = (result_df[temp_columns[4]] > 60).astype(int)
    result_df['Temp66'] = (result_df[temp_columns[5]] > 80).astype(int)
    result_df['TempSum'] = result_df[['Temp11', 'Temp22', 'Temp33', 'Temp44', 'Temp55', 'Temp66']].sum(axis=1)

    return compiled_df, filtered_df, max_df, result_df

def create_excel(compiled_df, filtered_df, max_df, result_df):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Compiled Data"
    ws2 = wb.create_sheet("Filtered Data")
    ws3 = wb.create_sheet("Max Data")
    ws4 = wb.create_sheet("Result Data")

    def write_df_to_sheet(ws, df):
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

    write_df_to_sheet(ws1, compiled_df)
    write_df_to_sheet(ws2, filtered_df)
    write_df_to_sheet(ws3, max_df)
    write_df_to_sheet(ws4, result_df)

    # Header formatting
    header_fill = PatternFill(start_color='157B8F', end_color='157B8F', fill_type='solid')
    bold_font = Font(bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    for cell in ws4[1]:
        cell.fill = header_fill
        cell.font = bold_font
        cell.border = thin_border

    highlight_fill_90 = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    highlight_fill_80 = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    highlight_fill_60 = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    highlight_fill_neg = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    green_fill = PatternFill(start_color='00A400', end_color='00A400', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

    def highlight_column_a(ws, row, fill):
        ws.cell(row=row, column=1).fill = fill

    col_names = [cell.value for cell in ws4[1]]

    for row in ws4.iter_rows(min_row=2, max_row=ws4.max_row, min_col=1, max_col=ws4.max_column):
        for cell in row:
            col_name = col_names[cell.column - 1]
            val = cell.value

            if isinstance(val, (int, float)):
                if col_name in temp_columns[:4] and val > 90:
                    cell.fill = highlight_fill_90
                elif col_name == 'OilSumpTemp' and val > 80:
                    cell.fill = highlight_fill_80
                elif col_name == 'MeasuredTemperatureofrotorbearing' and val > 60:
                    cell.fill = highlight_fill_60
                elif col_name.startswith('Temp') and 'TempSum' not in col_name and val > 0:
                    cell.fill = highlight_fill_neg

                if col_name == 'TempSum':
                    if val == 0:
                        cell.fill = green_fill
                        highlight_column_a(ws4, cell.row, green_fill)
                    elif val == 1:
                        cell.fill = yellow_fill
                        highlight_column_a(ws4, cell.row, yellow_fill)
                    elif val > 1:
                        cell.fill = red_fill
                        highlight_column_a(ws4, cell.row, red_fill)

    def apply_heatmap(ws, header_row=1, start_row=2):
        headers = [cell.value for cell in ws[header_row]]
        for col in temp_columns:
            if col in headers:
                idx = headers.index(col) + 1
                col_letter = get_column_letter(idx)
                rule = ColorScaleRule(
                    start_type='min', start_color='63BE7B',
                    mid_type='percentile', mid_value=50, mid_color='FFEB84',
                    end_type='max', end_color='F8696B'
                )
                ws.conditional_formatting.add(f"{col_letter}{start_row}:{col_letter}{ws.max_row}", rule)

    apply_heatmap(ws4)

    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer

def plot_exceedance_charts_plotly(compiled_df):
    charts = {}
    for asset, group in compiled_df.groupby('Asset Name'):
        exceeded_cols = []
        for col, limit in thresholds.items():
            if col in group.columns and (group[col] > limit).any():
                exceeded_cols.append(col)

        if not exceeded_cols:
            continue

        # Melt data for Plotly (long format)
        melted_df = group.melt(
            id_vars=["Date"],
            value_vars=exceeded_cols,
            var_name="Metric",
            value_name="Value"
        )

        # Add column for limit lines
        melted_df["Limit"] = melted_df["Metric"].map(thresholds)

        # Build interactive plot
        fig = px.line(
            melted_df,
            x="Date",
            y="Value",
            color="Metric",
            title=f"📈 Temperature Exceedance for {asset}",
            template="plotly_dark",
            markers=True
        )

        # Add limit lines
        for metric in exceeded_cols:
            limit = thresholds[metric]
            fig.add_hline(
                y=limit,
                line_dash="dash",
                line_color="white",
                annotation_text=f"{metric} Limit: {limit}°C",
                annotation_position="top right"
            )

        fig.update_layout(
            xaxis_title="Date",
            yaxis_title="Temperature (°C)",
            hovermode="x unified"
        )

        charts[asset] = fig

    return charts

# === Streamlit UI ===

uploaded_files = st.file_uploader("Upload CSV files", accept_multiple_files=True, type='csv')

if uploaded_files:
    # Save uploaded files to temp dir (in-memory)
    tmp_files = []
    for uploaded_file in uploaded_files:
        tmp_files.append(uploaded_file)

    # Process files
    compiled_df, filtered_df, max_df, result_df = process_data(tmp_files)

    if compiled_df is not None:
        st.subheader("🔍 Filter Options")

        # --- Asset Name Filter ---
        asset_names = compiled_df['Asset Name'].dropna().unique().tolist()
        selected_assets = st.multiselect("Select Asset(s)", asset_names, default=asset_names)

        # --- Temperature Column Filter ---
        available_temp_cols = [col for col in temp_columns if col in compiled_df.columns]
        selected_temp_cols = st.multiselect("Select Temperature Columns", available_temp_cols, default=available_temp_cols)

        # --- Date Range Filter ---
        try:
            date_min = compiled_df['Date'].min().date()
            date_max = compiled_df['Date'].max().date()
            date_range = st.date_input("Select Date Range", [date_min, date_max])
        except Exception as e:
            st.warning("Date column error. Check your 'Date' format.")
            st.stop()

        # === Apply Filters ===
        filtered_view_df = compiled_df.copy()

        if selected_assets:
            filtered_view_df = filtered_view_df[filtered_view_df['Asset Name'].isin(selected_assets)]
            result_df = result_df[result_df['Asset Name'].isin(selected_assets)]

        if len(date_range) == 2:
            start_date = pd.to_datetime(date_range[0])
            end_date = pd.to_datetime(date_range[1])
            filtered_view_df = filtered_view_df[
                (filtered_view_df['Date'] >= start_date) &
                (filtered_view_df['Date'] <= end_date)
            ]

        if filtered_view_df.empty:
            st.warning("⚠️ No data matching selected filters.")
        else:
            st.success(f"✅ Showing data for {len(filtered_view_df)} rows.")

            # Show result table
            st.subheader("📋 Result Data with Flags")
            st.dataframe(result_df)

            # Excel download
            excel_buffer = create_excel(compiled_df, filtered_df, max_df, result_df)
            st.download_button(
                label="Download Excel Report",
                data=excel_buffer,
                file_name="final_report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            # Filtered Charts
            st.subheader("📈 Temperature Exceedance Charts")

            # Only plot for filtered assets
            charts = plot_exceedance_charts_plotly(filtered_view_df)

            if not charts:
                st.info("No temperature exceedance detected for selected filters.")
            else:
                for asset, fig in charts.items():
                    st.markdown(f"**{asset}**")
                    st.plotly_chart(fig, use_container_width=True)

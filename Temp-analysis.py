import pandas as pd
import glob
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from datetime import datetime, timedelta

# Get yesterday's date
yesterday = datetime.now() - timedelta(days=1)
yesterday_str = yesterday.strftime('%Y-%m-%d')  # Format: 2025-08-20

# === SETUP ===
folder_path = r'D:\Temp'  # Folder containing CSV files
output_file_path = fr'D:\Isai\Suzlon\On_ {yesterday_str} _Suzlon.xlsx'  # Final Excel output
active_power_threshold = 500

# === Define Temperature Columns ===
temp_columns = [
    'OilSumpTemp',
    'GearboxHighSpeedShaftDrivenEndtemp',
    'GearboxHighSpeedShaftNonDrivenEndtemp',
    'Temperaturemeasurementforgeneratorbearingdriveend',
    'Temperaturemeasurementforgeneratorbearingnondriveend',
    'MeasuredTemperatureofrotorbearing'
]

required_cols = temp_columns + ['Asset Name', 'ActivepowerGeneration']

# === Step 1: Read and Combine All CSVs ===
csv_files = glob.glob(os.path.join(folder_path, '*.csv'))
raw_dfs = []

for file in csv_files:
    try:
        df = pd.read_csv(file)
        if 'Date' in df.columns:
            df['Date'] = pd.to_datetime(df['Date'], format='%d-%m-%Y %H:%M:%S', errors='coerce')
        if not df.empty:
            raw_dfs.append(df)
    except Exception as e:
        print(f"Error reading {file}: {e}")

if not raw_dfs:
    raise ValueError("No valid CSV files loaded.")

compiled_df = pd.concat(raw_dfs, ignore_index=True)

# === Step 2: Filter Data ===
missing = [col for col in required_cols if col not in compiled_df.columns]
if missing:
    raise KeyError(f"Missing required columns: {missing}")

filtered_df = compiled_df[(compiled_df['ActivepowerGeneration'] > 0)
]

# === Step 3: Max Aggregation ===
max_df = filtered_df.groupby('Asset Name')[temp_columns + ['ActivepowerGeneration']].max().reset_index()

# === Step 4: Add Temp Flags & TempSum ===
result_df = max_df.copy()

result_df['Temp11'] = (result_df[temp_columns[0]] >= 80).astype(int)
result_df['Temp22'] = (result_df[temp_columns[1]] >= 90).astype(int)
result_df['Temp33'] = (result_df[temp_columns[2]] >= 90).astype(int)
result_df['Temp44'] = (result_df[temp_columns[3]] >= 90).astype(int)
result_df['Temp55'] = (result_df[temp_columns[4]] >= 90).astype(int)
result_df['Temp66'] = (result_df[temp_columns[5]] >= 60).astype(int)

result_df['TempSum'] = result_df[['Temp11', 'Temp22', 'Temp33', 'Temp44', 'Temp55', 'Temp66']].sum(axis=1)

# === Step 5: Create Excel Workbook ===
wb = Workbook()
ws1 = wb.active
ws1.title = "Compiled Data"
ws2 = wb.create_sheet("Filtered Data")
ws3 = wb.create_sheet("Max Data")
ws4 = wb.create_sheet("Result Data")

# Helper to write DataFrame to worksheet
def write_df_to_sheet(ws, df):
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

# Write data to sheets
write_df_to_sheet(ws1, compiled_df)
write_df_to_sheet(ws2, filtered_df)
write_df_to_sheet(ws3, max_df)
write_df_to_sheet(ws4, result_df)

# === Apply Header Styles in Result Sheet ===
header_fill = PatternFill(start_color='157B8F', end_color='157B8F', fill_type='solid')
bold_font = Font(bold=True, color='000000')
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))

for cell in ws4[1]:
    cell.fill = header_fill
    cell.font = bold_font
    cell.border = thin_border

# === Conditional Formatting ===
highlight_fill_90 = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # pink >90
highlight_fill_80 = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # yellow >80
highlight_fill_60 = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # green >60
highlight_fill_neg = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") # yellow for <0
green_fill = PatternFill(start_color='00A400', end_color='00A400', fill_type='solid')
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

# Highlight column A based on TempSum
def highlight_column_a(ws, row, fill):
    ws.cell(row=row, column=1).fill = fill

# Get headers
col_names = [cell.value for cell in ws4[1]]

for row in ws4.iter_rows(min_row=2, max_row=ws4.max_row, min_col=1, max_col=ws4.max_column):
    for cell in row:
        col_name = col_names[cell.column - 1]
        val = cell.value

        if isinstance(val, (int, float)):
            if col_name in temp_columns[:4] and val >= 90:
                cell.fill = highlight_fill_90
            elif col_name == 'OilSumpTemp' and val >= 80:
                cell.fill = highlight_fill_80
            elif col_name == 'MeasuredTemperatureofrotorbearing' and val >= 60:
                cell.fill = highlight_fill_60
            elif col_name.startswith('Temp') and 'TempSum' not in col_name and val > 0:
                cell.fill = highlight_fill_neg

            if col_name == 'TempSum':
                if val == 0:
                    cell.fill = green_fill
                    highlight_column_a(ws4, cell.row, green_fill)
                elif val == 1:
                    cell.fill = yellow_fill
                    highlight_column_a(ws4, cell.row, yellow_fill)
                elif val > 1:
                    cell.fill = red_fill
                    highlight_column_a(ws4, cell.row, red_fill)

# === Apply Heatmap (Color Gradient) ===
def apply_heatmap_to_temp_columns(ws, header_row=1, start_row=2):
    headers = [cell.value for cell in ws[header_row]]
    for col_name in temp_columns:
        if col_name in headers:
            col_idx = headers.index(col_name) + 1
            col_letter = get_column_letter(col_idx)
            heatmap_rule = ColorScaleRule(
                start_type='min', start_color='63BE7B',  # Green
                mid_type='percentile', mid_value=50, mid_color='FFEB84',  # Yellow
                end_type='max', end_color='F8696B'  # Red
            )
            ws.conditional_formatting.add(f"{col_letter}{start_row}:{col_letter}{ws.max_row}", heatmap_rule)

apply_heatmap_to_temp_columns(ws4)

# === Save Workbook ===
wb.save(output_file_path)
print(f"\n✅ Excel report saved with heatmap and formatting at:\n{output_file_path}")
